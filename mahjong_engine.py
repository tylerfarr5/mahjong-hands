"""
Mahjong Hand Combination Generator (v2)
=========================================
Declarative engine: describe a hand as a list of "groups" (pungs, kongs,
quints, pairs, singles, runs, dragon/wind/flower sets) plus constraints on
suit variables and (optionally) shared value variables, and it enumerates
every physically-valid way to build that hand from the 152-tile set.

Physical constraints always enforced:
  - max 4 real copies of any single numbered/dragon/wind tile identity
  - max 8 flowers total
  - max 8 jokers total across the whole hand
  - jokers may only fill groups of size >= 3 (never singles/pairs); a
    pung/kong/quint/sextet may be built entirely from jokers if desired
  - flower groups can use up to 8 real copies (vs. 4 for everything else)
    since there are 8 flowers in the set

--------------------------------------------------------------------------
GROUP FIELDS
--------------------------------------------------------------------------
  kind        : 'same'              -> N copies of one tile identity
                'complement_singles' -> one single of every value in a
                                        domain EXCEPT the resolved value
                                        of a given value_var (used for
                                        "pair is any of X, the rest are
                                        singles" style hands)
  size        : tile count (for 'same')
  tile_type   : 'number' | 'dragon' | 'wind' | 'flower'
  value       : (tile_type='number' only) one of:
                   - int                      literal value
                   - [int, int, ...]          branch/OR over these values
                   - 'VarName'                shared value, resolved via
                                               hand_spec['value_vars']
                   - ('offset', 'VarName', k)  resolved = value_vars[Var]+k
                                               (for consecutive runs)
  suit        : (tile_type='number'/'dragon' only) one of:
                   - None/omitted             (n/a - wind & flower only)
                   - 'Crak'|'Bam'|'Dot'        literal suit
                   - 'A' (etc.)                a suit variable name,
                                               resolved via suit_constraints
                   - ['A','B','C']             branch/OR: use whichever
                                               variable's resolved suit
  wind        : 'North'|'East'|'South'|'West' (tile_type='wind' only)
  allow_joker : bool override (default: size>=3, and always False for
                tile_type='flower')

  -- complement_singles only --
  domain      : list of ints, e.g. [1,3,5,7,9]
  value_var   : name of the value_var whose resolved value is EXCLUDED
  (tile_type/suit apply as above; always size 1 each, never joker)

--------------------------------------------------------------------------
HAND-LEVEL FIELDS
--------------------------------------------------------------------------
  hand_id           : label shown in output column A
  category          : optional, used for sheet grouping/naming
  groups            : list of group dicts (above)
  suit_constraints  : {'vars': ['A','B',...], 'distinct_count': k}
                        k = None  -> no constraint between vars
                        k = len(vars) -> all pairwise distinct (default
                                         if 'distinct_count' omitted)
                        k = m (m < len(vars)) -> enumerate every way the
                                         vars could take on exactly m
                                         distinct suits (handles hands
                                         where more label-slots exist
                                         than actual distinct suits used)
  value_vars        : {'VarName': [domain ints]}  e.g. {'V': [1,3,5,7,9]}
                        each combination of value_vars is tried; groups
                        reference them by name (see 'value' above)
"""

import itertools
import openpyxl
from openpyxl.styles import Font, Alignment

SUITS = ['Crak', 'Bam', 'Dot']
WINDS = ['North', 'East', 'South', 'West']
MAX_COPIES = 4
TOTAL_JOKERS = 8
TOTAL_FLOWERS = 8

COLUMN_HEADER_1 = (
    ['Suit'] + ['Crak'] * 10 + ['Bam'] * 10 + ['Dot'] * 10 +
    ['Flower'] + ['Wind'] * 4 + ['Joker']
)
COLUMN_HEADER_2 = (
    ['Value'] +
    [1,2,3,4,5,6,7,8,9,'Red Dragon'] +
    [1,2,3,4,5,6,7,8,9,'Green Dragon'] +
    [1,2,3,4,5,6,7,8,9,'White Dragon'] +
    ['Flower','North','East','South','West','Joker']
)
NCOLS = len(COLUMN_HEADER_1)  # 37

SUIT_BASE = {'Crak': 1, 'Bam': 11, 'Dot': 21}
SUIT_DRAGON_COL = {'Crak': 10, 'Bam': 20, 'Dot': 30}
WIND_COL = {'North': 32, 'East': 33, 'South': 34, 'West': 35}
FLOWER_COL = 31
JOKER_COL = 36


def _col(tile_type, suit=None, value=None, wind=None):
    if tile_type == 'number':
        return SUIT_BASE[suit] + (value - 1)
    if tile_type == 'dragon':
        return SUIT_DRAGON_COL[suit]
    if tile_type == 'wind':
        return WIND_COL[wind]
    if tile_type == 'flower':
        return FLOWER_COL
    raise ValueError(f"Unknown tile_type {tile_type}")


def _col_cap(col):
    return TOTAL_FLOWERS if col == FLOWER_COL else MAX_COPIES


def _real_joker_options(size, allow_joker, max_real=MAX_COPIES):
    if not allow_joker:
        return [(size, 0)]
    # real can go all the way to 0 - an entire pung/kong/quint/sextet
    # may be built from jokers alone
    return [(r, size - r) for r in range(0, min(max_real, size) + 1)]


def _resolve_suit_var_assignments(suit_constraints):
    variables = suit_constraints.get('vars', []) if suit_constraints else []
    if not variables:
        yield {}
        return
    n = len(variables)
    default_k = n
    k = suit_constraints.get('distinct_count', default_k)
    for combo in itertools.product(SUITS, repeat=n):
        if k is None or len(set(combo)) == k:
            yield dict(zip(variables, combo))


def _resolve_value_var_assignments(value_vars):
    if not value_vars:
        yield {}
        return
    names = list(value_vars.keys())
    domains = [value_vars[n] for n in names]
    for combo in itertools.product(*domains):
        yield dict(zip(names, combo))


def _resolve_suit_field(suit_field, suit_map):
    """Return list of candidate literal suits for this group."""
    if suit_field is None:
        return [None]
    candidates = suit_field if isinstance(suit_field, list) else [suit_field]
    out = []
    for c in candidates:
        if c in SUITS:
            out.append(c)
        elif c in suit_map:
            out.append(suit_map[c])
        else:
            raise ValueError(f"Unresolved suit reference: {c}")
    return out


def _resolve_value_field(value_field, value_map):
    """Return list of candidate literal values for this group."""
    if value_field is None:
        return [None]
    if isinstance(value_field, tuple) and value_field[0] == 'offset':
        _, var, k = value_field
        return [value_map[var] + k]
    if isinstance(value_field, list):
        return value_field
    if isinstance(value_field, str):
        return [value_map[value_field]]
    return [value_field]  # literal int


def _group_choices(g, suit_map, value_map):
    """
    Returns list of "placements". Each placement is a list of
    (col, real, joker) tuples to add simultaneously to the row.
    """
    if g['kind'] == 'complement_singles':
        domain = g['domain']
        excl = value_map[g['value_var']]
        suit_candidates = _resolve_suit_field(g.get('suit'), suit_map)
        placements = []
        for suit in suit_candidates:
            placement = []
            for v in domain:
                if v == excl:
                    continue
                col = _col(g['tile_type'], suit=suit, value=v)
                placement.append((col, 1, 0))
            placements.append(placement)
        return placements

    if g['kind'] == 'same':
        tile_type = g['tile_type']
        size = g['size']
        allow_joker = g.get('allow_joker', size >= 3)
        max_real = TOTAL_FLOWERS if tile_type == 'flower' else MAX_COPIES
        rj_opts = _real_joker_options(size, allow_joker, max_real)
        suit_candidates = _resolve_suit_field(g.get('suit'), suit_map)
        value_candidates = _resolve_value_field(g.get('value'), value_map)
        wind_field = g.get('wind')
        wind_candidates = wind_field if isinstance(wind_field, list) else [wind_field]
        placements = []
        for suit in suit_candidates:
            for value in value_candidates:
                for wind in wind_candidates:
                    for (real, joker) in rj_opts:
                        col = _col(tile_type, suit=suit, value=value, wind=wind)
                        placements.append([(col, real, joker)])
        return placements

    raise ValueError(f"Unknown group kind {g['kind']}")


def generate_combinations(hand_spec, total_jokers=TOTAL_JOKERS, verbose=True):
    groups = hand_spec['groups']
    suit_constraints = hand_spec.get('suit_constraints', {})
    value_vars = hand_spec.get('value_vars', {})
    hand_id = hand_spec['hand_id']

    # sanity check: declared sizes should sum to 14 (skip if complement
    # groups involved, since their effective size is domain-dependent but
    # constant across resolutions - compute using first domain choice)
    total_size = 0
    for g in groups:
        if g['kind'] == 'same':
            total_size += g['size']
        elif g['kind'] == 'complement_singles':
            total_size += len(g['domain']) - 1
    if total_size != 14 and verbose:
        print(f"  [!] WARNING: hand {hand_id} groups sum to {total_size}, not 14")

    rows = []
    for suit_map in _resolve_suit_var_assignments(suit_constraints):
        for value_map in _resolve_value_var_assignments(value_vars):
            per_group_choices = [_group_choices(g, suit_map, value_map) for g in groups]
            for combo in itertools.product(*per_group_choices):
                row = [0] * NCOLS
                row[0] = hand_id
                total_joker = 0
                for placement in combo:
                    for (col, real, joker) in placement:
                        row[col] += real
                        total_joker += joker
                if total_joker > total_jokers:
                    continue
                valid = True
                for col in range(1, NCOLS - 1):
                    if row[col] > _col_cap(col):
                        valid = False
                        break
                if not valid:
                    continue
                row[JOKER_COL] = total_joker
                rows.append(row)

    # de-duplicate (different suit_map/value_map resolutions can sometimes
    # coincide in the final tile counts, e.g. symmetric groups)
    seen = set()
    unique_rows = []
    for r in rows:
        key = tuple(r)
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)

    if verbose:
        print(f"Hand {hand_id}: {len(unique_rows)} combinations")
    return unique_rows


def write_workbook(hands, output_path):
    """hands: list of (hand_spec, rows) or hand_spec (auto-generates rows)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    used_titles = set()

    for item in hands:
        if isinstance(item, tuple):
            hand_spec, rows = item
        else:
            hand_spec = item
            rows = generate_combinations(hand_spec)

        raw_title = str(hand_spec['hand_id'])
        for ch in [':', '\\', '/', '?', '*', '[', ']']:
            raw_title = raw_title.replace(ch, '-')
        title = raw_title[:31]
        base_title = title
        i = 2
        while title in used_titles:
            suffix = f"_{i}"
            title = base_title[:31 - len(suffix)] + suffix
            i += 1
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        ws.append(COLUMN_HEADER_1)
        ws.append(COLUMN_HEADER_2)
        for r in rows:
            ws.append(r)
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = bold
                cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = "B3"

    wb.save(output_path)
    print("Saved:", output_path)
