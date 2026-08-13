import sys
sys.path.insert(0, '/home/claude/mahjong')
from mahjong_engine import (generate_combinations, COLUMN_HEADER_1, COLUMN_HEADER_2,
                             NCOLS, SUIT_BASE, SUIT_DRAGON_COL, WIND_COL, FLOWER_COL, JOKER_COL)
from all_hands import hands  # 72 hand specs, in stable order

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

BOLD = Font(bold=True)
CENTER = Alignment(horizontal='center')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')   # light yellow - cells to edit
CALC_FONT = Font(color='808080', italic=True)          # helper/engine cells
HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')

# 36 tile labels, in the exact column order used everywhere (matches
# mahjong_engine's tile columns, i.e. COLUMN_HEADER_2[1:])
TILE_LABELS = []
for h1, h2 in zip(COLUMN_HEADER_1[1:], COLUMN_HEADER_2[1:]):
    if h1 in ('Crak', 'Bam', 'Dot'):
        TILE_LABELS.append(f"{h1} {h2}" if isinstance(h2, int) else f"{h1} {h2}")
    else:
        TILE_LABELS.append(str(h2))
assert len(TILE_LABELS) == 36

# Example pre-fill (matches the user's own earlier example: 13 tiles)
EXAMPLE = {
    'Crak 1': 2, 'Crak 3': 1, 'Crak 5': 1, 'Crak 7': 1, 'Crak 9': 1,
    'Bam 3': 3, 'Dot 3': 2, 'Joker': 2,
}

wb = openpyxl.Workbook()
wb.remove(wb.active)

# =====================================================================
# 1) Generate all combinations once (used for All Combos + per-hand sheets)
# =====================================================================
hand_rows = []
for idx, h in enumerate(hands, start=1):
    rows = generate_combinations(h, verbose=False)
    hand_rows.append((idx, h, rows))
print("Generated all hands:", sum(len(r) for _, _, r in hand_rows), "total rows")

MAX_DATA_ROWS = 100000  # safe upper bound used in every formula range below

# =====================================================================
# 2) All Combos sheet: HandName | 36 tile cols | Overlap | HandID | Key
# =====================================================================
ws_all = wb.create_sheet('All Combos')
header = ['Hand'] + TILE_LABELS + ['Overlap', 'HandID', 'Key']
ws_all.append(header)
for cell in ws_all[1]:
    cell.font = BOLD
    cell.fill = HEADER_FILL

r = 2
for idx, h, rows in hand_rows:
    for row in rows:
        rr = r
        overlap_formula = (f"=SUMPRODUCT((B{rr}:AK{rr}+'Hand Profile'!$C$20:$AL$20"
                            f"-ABS(B{rr}:AK{rr}-'Hand Profile'!$C$20:$AL$20)))/2")
        key_formula = f"=AM{rr}*100+AL{rr}"
        full_row = row + [overlap_formula, idx, key_formula]
        ws_all.append(full_row)
        r += 1
ws_all.freeze_panes = "B2"
print("All Combos rows written:", ws_all.max_row)

# =====================================================================
# 3) Top Combos sheet (informational: # of combinations per hand)
# =====================================================================
ws_top = wb.create_sheet('Top Combos')
ws_top.append(['Hand', 'Combinations Available'])
for cell in ws_top[1]:
    cell.font = BOLD
    cell.fill = HEADER_FILL
for idx, h, rows in hand_rows:
    ws_top.append([h['hand_id'], len(rows)])
ws_top.column_dimensions['A'].width = 32

# =====================================================================
# 4) Hand Profile sheet
# =====================================================================
ws = wb.create_sheet('Hand Profile', 0)  # first sheet
ws.sheet_view.showGridLines = False

ws['A1'] = 'MAHJONG HAND PROFILE & RECOMMENDATIONS'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'Enter the 13 tiles you are holding in the yellow cells below.'
ws['A2'].font = Font(italic=True, size=10)

def block(start_row, suit_label, values, count_prefill=None):
    """Writes a Suit/Value/Count 3-row input block starting at start_row, col C."""
    count_prefill = count_prefill or {}
    n = len(values)
    ws.cell(row=start_row, column=3, value='Suit').font = BOLD
    ws.cell(row=start_row + 1, column=3, value='Value').font = BOLD
    ws.cell(row=start_row + 2, column=3, value='Count').font = BOLD
    for i, v in enumerate(values):
        col = 4 + i
        ws.cell(row=start_row, column=col, value=suit_label).font = BOLD
        ws.cell(row=start_row, column=col).alignment = CENTER
        ws.cell(row=start_row + 1, column=col, value=v).font = BOLD
        ws.cell(row=start_row + 1, column=col).alignment = CENTER
        c = ws.cell(row=start_row + 2, column=col)
        label = f"{suit_label} {v}" if suit_label else str(v)
        c.value = count_prefill.get(label)
        c.fill = INPUT_FILL
        c.alignment = CENTER

block(3, 'Crak', [1, 2, 3, 4, 5, 6, 7, 8, 9, 'Red Dragon'], EXAMPLE)
block(7, 'Bam', [1, 2, 3, 4, 5, 6, 7, 8, 9, 'Green Dragon'], EXAMPLE)
block(11, 'Dot', [1, 2, 3, 4, 5, 6, 7, 8, 9, 'White Dragon'], EXAMPLE)

# Honors block (Flower, 4 winds, Joker) - 6 slots, custom labels since suit varies per cell
ws.cell(row=15, column=3, value='Suit').font = BOLD
ws.cell(row=16, column=3, value='Value').font = BOLD
ws.cell(row=17, column=3, value='Count').font = BOLD
honor_defs = [('Flower', 'Flower'), ('Wind', 'North'), ('Wind', 'East'),
              ('Wind', 'South'), ('Wind', 'West'), ('Joker', 'Joker')]
for i, (suit_lbl, val_lbl) in enumerate(honor_defs):
    col = 4 + i
    ws.cell(row=15, column=col, value=suit_lbl).font = BOLD
    ws.cell(row=15, column=col).alignment = CENTER
    ws.cell(row=16, column=col, value=val_lbl).font = BOLD
    ws.cell(row=16, column=col).alignment = CENTER
    c = ws.cell(row=17, column=col)
    c.value = EXAMPLE.get(val_lbl)
    c.fill = INPUT_FILL
    c.alignment = CENTER

# Total tiles check
ws.cell(row=15, column=11, value='Total Tiles').font = BOLD
ws.cell(row=17, column=11, value='=SUM(D5:M5)+SUM(D9:M9)+SUM(D13:M13)+SUM(D17:I17)')
ws.cell(row=17, column=11).font = Font(bold=True, size=12)
ws.conditional_formatting.add(
    'K17',
    CellIsRule(operator='notEqual', formula=['13'], fill=PatternFill('solid', fgColor='FFC7CE'))
)
ws.cell(row=18, column=11, value='(should equal 13)').font = Font(italic=True, size=9, color='808080')

# Reference vector: 36 labels (row19) + 36 live values (row20), aligned to All Combos columns
ws.cell(row=19, column=2, value='Reference vector (auto-built from above; used by formulas - do not edit)').font = CALC_FONT
src_cells = (
    [f"{get_column_letter(4+i)}5" for i in range(10)] +   # Crak
    [f"{get_column_letter(4+i)}9" for i in range(10)] +   # Bam
    [f"{get_column_letter(4+i)}13" for i in range(10)] +  # Dot
    [f"{get_column_letter(4+i)}17" for i in range(6)]     # Flower, N, E, S, W, Joker
)
for i, (label, src) in enumerate(zip(TILE_LABELS, src_cells)):
    col = 3 + i  # C.. AL
    lc = ws.cell(row=19, column=col, value=label)
    lc.font = CALC_FONT
    lc.alignment = CENTER
    vc = ws.cell(row=20, column=col, value=f"=IF({src}=\"\",0,{src})")
    vc.font = CALC_FONT
    vc.alignment = CENTER
ws.row_dimensions[19].hidden = False

# =====================================================================
# 5) Full 72-hand scoring table (engine area), rows 23..(22+72)
# =====================================================================
TABLE_HEADER_ROW = 23
TABLE_START = 24
n_hands = len(hands)
TABLE_END = TABLE_START + n_hands - 1

ws.cell(row=TABLE_HEADER_ROW, column=3, value='ID').font = BOLD
ws.cell(row=TABLE_HEADER_ROW, column=4, value='Hand').font = BOLD
ws.cell(row=TABLE_HEADER_ROW, column=5, value='Tiles Matched').font = BOLD
ws.cell(row=TABLE_HEADER_ROW, column=6, value='Tiles Needed').font = BOLD
ws.cell(row=TABLE_HEADER_ROW, column=7, value='Match %').font = BOLD
ws.cell(row=TABLE_HEADER_ROW, column=8, value='Rank Key').font = CALC_FONT
ws.cell(row=TABLE_HEADER_ROW, column=9, value='Rank').font = CALC_FONT
for c in range(3, 10):
    ws.cell(row=TABLE_HEADER_ROW, column=c).fill = HEADER_FILL

for i, h in enumerate(hands):
    row = TABLE_START + i
    hid = i + 1
    ws.cell(row=row, column=3, value=hid)
    ws.cell(row=row, column=4, value=h['hand_id'])
    ws.cell(row=row, column=5,
            value=f"=_xlfn.MAXIFS('All Combos'!$AL$2:$AL${MAX_DATA_ROWS},"
                  f"'All Combos'!$AM$2:$AM${MAX_DATA_ROWS},C{row})")
    ws.cell(row=row, column=6, value=f"=14-E{row}")
    e_cell = f"E{row}"
    p_cell = ws.cell(row=row, column=7, value=f"={e_cell}/13")
    p_cell.number_format = '0.0%'
    ws.cell(row=row, column=8, value=f"={e_cell}*1000+(100000-ROW())*0.00001")
    ws.cell(row=row, column=9, value=f"=RANK({'H'}{row},$H${TABLE_START}:$H${TABLE_END},0)")

# =====================================================================
# 6) Top 5 recommendations (prominent, top-right of sheet)
# =====================================================================
ws.cell(row=2, column=14, value='TOP HAND RECOMMENDATIONS').font = Font(bold=True, size=14)
top_header_row = 3
headers = ['Rank', 'Hand', 'Tiles Matched (of 13)', 'Tiles Still Needed', 'Match %', "What's Needed (example)"]
for j, htext in enumerate(headers):
    c = ws.cell(row=top_header_row, column=14 + j, value=htext)
    c.font = BOLD
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, horizontal='center')

TOP_N = 5
DETAIL_START_ROW = 60  # per-rank 36-wide shortfall detail rows (hidden-ish, below)

for k in range(1, TOP_N + 1):
    row = top_header_row + k
    ws.cell(row=row, column=14, value=k).font = BOLD
    # HandID for this rank
    hid_formula = f"=INDEX($C${TABLE_START}:$C${TABLE_END},MATCH({k},$I${TABLE_START}:$I${TABLE_END},0))"
    ws.cell(row=row, column=15,
            value=f"=INDEX($D${TABLE_START}:$D${TABLE_END},MATCH({k},$I${TABLE_START}:$I${TABLE_END},0))")
    matched_cell = ws.cell(row=row, column=16,
            value=f"=INDEX($E${TABLE_START}:$E${TABLE_END},MATCH({k},$I${TABLE_START}:$I${TABLE_END},0))")
    ws.cell(row=row, column=17, value=f"=14-P{row}")
    pct_cell = ws.cell(row=row, column=18, value=f"=P{row}/13")
    pct_cell.number_format = '0.0%'

    # hidden helper columns: HandID (col T=20), RowIdx (col U=21)
    ws.cell(row=row, column=20, value=hid_formula)
    ws.cell(row=row, column=21,
            value=f"=MATCH(T{row}*100+P{row},'All Combos'!$AN$2:$AN${MAX_DATA_ROWS},0)+1")

    # 36-wide shortfall detail row for this rank
    drow = DETAIL_START_ROW + k
    for i in range(36):
        col = 3 + i  # C..AL, aligned with reference vector columns
        col_letter = get_column_letter(col)
        target_ref = f"INDEX('All Combos'!$B:$AK,$U{row},{i+1})"
        user_ref = f"{col_letter}$20"
        label_ref = f"{col_letter}$19"
        formula = (f'=IF({target_ref}>{user_ref},{label_ref}&" +"&({target_ref}-{user_ref}),"")')
        ws.cell(row=drow, column=col, value=formula).font = CALC_FONT

    detail_range = f"C{drow}:AL{drow}"
    needed_cell = ws.cell(row=row, column=19,
            value=f'=_xlfn.TEXTJOIN(", ",TRUE,{detail_range})')
    needed_cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.cell(row=59, column=3, value='Per-rank shortfall detail (auto; feeds the "What\'s Needed" column above)').font = CALC_FONT

# column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
for col in 'EFGHIJKLM':
    ws.column_dimensions[col].width = 8
ws.column_dimensions['N'].width = 6
ws.column_dimensions['O'].width = 32
ws.column_dimensions['P'].width = 12
ws.column_dimensions['Q'].width = 12
ws.column_dimensions['R'].width = 10
ws.column_dimensions['S'].width = 60
for k in range(1, TOP_N + 1):
    ws.row_dimensions[top_header_row + k].height = 45

# =====================================================================
# 7) Individual per-hand sheets (same as before)
# =====================================================================
used_titles = set()
for idx, h, rows in hand_rows:
    raw_title = str(h['hand_id'])
    for ch in [':', '\\', '/', '?', '*', '[', ']']:
        raw_title = raw_title.replace(ch, '-')
    title = raw_title[:31]
    base_title = title
    i = 2
    while title in used_titles:
        suffix = f"_{i}"
        title = base_title[:31 - len(suffix)] + suffix
        i += 1
    used_titles.add(title)

    wsh = wb.create_sheet(title=title)
    wsh.append(COLUMN_HEADER_1)
    wsh.append(COLUMN_HEADER_2)
    for row in rows:
        wsh.append(row)
    for rowcells in wsh.iter_rows(min_row=1, max_row=2):
        for cell in rowcells:
            cell.font = BOLD
            cell.alignment = CENTER
    wsh.freeze_panes = "B3"

out_path = '/mnt/user-data/outputs/Mahjong_Hand_Recommender.xlsx'
wb.save(out_path)
print("Saved:", out_path)
